import openpyxl

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

x = [0, 55, 100, 170, 313, 470, 640, 820]

# Write data to cells
ws['A1'] = 'Denumirea'
ws['B1'] = x[0]
ws['C1'] = x[1]
ws['D1'] = x[2]
ws['E1'] = x[3]
ws['F1'] = x[4]
ws['G1'] = x[5]
ws['H1'] = x[6]
ws['I1'] = x[7]

# Save the workbook
wb.save('column_sizes.xlsx')